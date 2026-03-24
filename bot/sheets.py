# -*- coding: utf-8 -*-
"""
Модуль работы с таблицами — сохранение заказов в Excel на Яндекс Диске.

Колонки: Имя, Контакт, Товар, Количество, Адрес, Оплата, Комментарий, Telegram ID

При ошибке Диска — заказ сохраняется в orders_backup.csv.
При каждом новом заказе пытаемся перенести накопленный backup в облако.
"""

import csv
import logging
import tempfile
import time
from pathlib import Path

import yadisk
from openpyxl import Workbook, load_workbook

from config import YANDEX_DISK_FILE_PATH, YANDEX_DISK_TOKEN

logger = logging.getLogger(__name__)

HEADERS = [
    "Имя", "Контакт", "Товар", "Количество", "Адрес", "Оплата", "Комментарий", "Telegram ID"
]

_BACKUP_CSV = Path(__file__).resolve().parent.parent / "orders_backup.csv"


def _get_client() -> yadisk.Client:
    return yadisk.Client(token=YANDEX_DISK_TOKEN)


def _ensure_parent_dir(client: yadisk.Client, file_path: str) -> None:
    parts = [p for p in file_path.split("/") if p]
    if len(parts) > 1:
        parent = "/" + "/".join(parts[:-1])
        if not client.exists(parent):
            client.makedirs(parent)


def _read_backup() -> list[list[str]]:
    """Читает заказы из резервного CSV. Возвращает список строк (без заголовка)."""
    if not _BACKUP_CSV.exists():
        return []
    rows = []
    with open(_BACKUP_CSV, "r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader, None)  # пропуск заголовка
        for row in reader:
            if len(row) >= 8:
                rows.append(row)
            elif row:
                rows.append(row + [""] * (8 - len(row)))
    return rows


def _clear_backup() -> None:
    """Очищает backup — оставляет только заголовок."""
    with open(_BACKUP_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
    logger.info("Резервный файл очищен после успешного переноса в облако")


def try_sync_backup() -> bool:
    """
    Пытается перенести заказы из backup в облако (когда applications.xlsx закрыт).
    Вызывается при старте бота и не требует нового заказа.
    Возвращает True при успешном переносе.
    """
    backup_rows = _read_backup()
    if not backup_rows:
        return False

    try:
        client = _get_client()
        client.check_token()
        disk_path = YANDEX_DISK_FILE_PATH
        if not disk_path.startswith("/"):
            disk_path = "/" + disk_path

        with tempfile.TemporaryDirectory() as tmpdir:
            local_path = Path(tmpdir) / "applications.xlsx"
            if client.exists(disk_path):
                client.download(disk_path, str(local_path))
                wb = load_workbook(local_path)
                ws = wb.active
                if ws.max_row == 0 or not any(cell.value for cell in ws[1]):
                    ws.append(HEADERS)
            else:
                _ensure_parent_dir(client, disk_path)
                wb = Workbook()
                ws = wb.active
                ws.title = "Заказы"
                ws.append(HEADERS)
            for row in backup_rows:
                ws.append(row)
            wb.save(local_path)
            _upload_with_retries(client, local_path, disk_path)

        _clear_backup()
        logger.info("При старте перенесено %s заказов из backup в облако", len(backup_rows))
        return True
    except Exception as e:
        logger.debug("Синхронизация backup при старте не удалась (файл может быть открыт): %s", e)
        return False


def _upload_with_retries(client: yadisk.Client, local_path: Path, disk_path: str) -> None:
    """Загружает файл с повторными попытками при блокировке."""
    for attempt in range(6):
        try:
            client.upload(str(local_path), disk_path, overwrite=True)
            return
        except yadisk.exceptions.ResourceIsLockedError:
            if attempt < 5:
                wait = 5 + attempt * 5
                logger.warning("Файл заблокирован, повтор через %s сек (попытка %s/5)", wait, attempt + 1)
                time.sleep(wait)
            else:
                raise


def append_order(
    name: str,
    contact: str,
    product: str,
    quantity: str,
    address: str,
    payment: str,
    comment: str = "",
    telegram_user_id: int | None = None,
) -> tuple[bool, bool]:
    """
    Добавляет заказ в Excel на Яндекс Диске.
    Сначала пытается перенести заказы из backup (если файл был закрыт).
    """
    new_row = [
        name, contact, product, quantity, address, payment, comment,
        str(telegram_user_id) if telegram_user_id else "",
    ]

    try:
        client = _get_client()
        client.check_token()

        disk_path = YANDEX_DISK_FILE_PATH
        if not disk_path.startswith("/"):
            disk_path = "/" + disk_path

        backup_rows = _read_backup()

        with tempfile.TemporaryDirectory() as tmpdir:
            local_path = Path(tmpdir) / "applications.xlsx"

            if client.exists(disk_path):
                client.download(disk_path, str(local_path))
                wb = load_workbook(local_path)
                ws = wb.active
                if ws.max_row == 0 or not any(cell.value for cell in ws[1]):
                    ws.append(HEADERS)
            else:
                _ensure_parent_dir(client, disk_path)
                wb = Workbook()
                ws = wb.active
                ws.title = "Заказы"
                ws.append(HEADERS)

            # Сначала добавляем заказы из backup (если были)
            for row in backup_rows:
                ws.append(row)
            # Затем новый заказ
            ws.append(new_row)
            wb.save(local_path)

            try:
                _upload_with_retries(client, local_path, disk_path)
            except yadisk.exceptions.ResourceIsLockedError:
                raise

        # Успех — очищаем backup
        if backup_rows:
            _clear_backup()
            logger.info("Перенесено %s заказов из backup + новый заказ в облако", len(backup_rows))
        else:
            logger.info("Заказ успешно сохранён в Яндекс Диск: %s", name)
        return (True, False)

    except yadisk.exceptions.UnauthorizedError as e:
        logger.error("Яндекс Диск: токен недействителен: %s", e)
        ok = _save_to_backup(name, contact, product, quantity, address, payment, comment, telegram_user_id)
        return (False, ok)
    except yadisk.exceptions.ForbiddenError as e:
        logger.error("Яндекс Диск: недостаточно прав: %s", e)
        ok = _save_to_backup(name, contact, product, quantity, address, payment, comment, telegram_user_id)
        return (False, ok)
    except yadisk.exceptions.ResourceIsLockedError as e:
        logger.error("Яндекс Диск: файл заблокирован: %s", e)
        ok = _save_to_backup(name, contact, product, quantity, address, payment, comment, telegram_user_id)
        return (False, ok)
    except Exception as e:
        logger.exception("Ошибка при сохранении заказа: %s", e)
        ok = _save_to_backup(name, contact, product, quantity, address, payment, comment, telegram_user_id)
        return (False, ok)


def _save_to_backup(
    name: str, contact: str, product: str, quantity: str, address: str,
    payment: str, comment: str, telegram_user_id: int | None,
) -> bool:
    try:
        write_header = not _BACKUP_CSV.exists()
        with open(_BACKUP_CSV, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if write_header:
                w.writerow(HEADERS)
            w.writerow([name, contact, product, quantity, address, payment, comment, str(telegram_user_id or "")])
        logger.info("Заказ сохранён в резервный файл: %s", _BACKUP_CSV)
        return True
    except Exception as ex:
        logger.exception("Не удалось сохранить в резерв: %s", ex)
        return False


