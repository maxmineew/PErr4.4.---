# -*- coding: utf-8 -*-
"""Запись заявок с сайта в Excel на Яндекс Диске. Лист «Заявки с сайта»."""

import logging
import tempfile
import time
from datetime import datetime
from pathlib import Path

import yadisk
from openpyxl import Workbook, load_workbook

from app.config import YANDEX_DISK_FILE_PATH, YANDEX_DISK_TOKEN

logger = logging.getLogger(__name__)

SHEET_NAME = "Заявки с сайта"
HEADERS = ["Имя", "Контакт", "Описание запроса", "Дата", "Источник"]


def _client() -> yadisk.Client:
    return yadisk.Client(token=YANDEX_DISK_TOKEN)


def _ensure_dir(client: yadisk.Client, path: str) -> None:
    parts = [p for p in path.split("/") if p]
    if len(parts) > 1:
        parent = "/" + "/".join(parts[:-1])
        if not client.exists(parent):
            client.makedirs(parent)


def _upload_retry(client: yadisk.Client, local: Path, remote: str) -> None:
    for i in range(5):
        try:
            client.upload(str(local), remote, overwrite=True)
            return
        except yadisk.exceptions.ResourceIsLockedError:
            if i < 4:
                time.sleep(5 + i * 5)
            else:
                raise


def append_lead(name: str, contact: str, description: str) -> bool:
    """Добавляет заявку с сайта в таблицу."""
    try:
        client = _client()
        client.check_token()
        disk_path = YANDEX_DISK_FILE_PATH if YANDEX_DISK_FILE_PATH.startswith("/") else "/" + YANDEX_DISK_FILE_PATH

        with tempfile.TemporaryDirectory() as tmp:
            lp = Path(tmp) / "web_leads.xlsx"
            if client.exists(disk_path):
                client.download(disk_path, str(lp))
                wb = load_workbook(lp)
                if SHEET_NAME in wb.sheetnames:
                    ws = wb[SHEET_NAME]
                else:
                    ws = wb.create_sheet(SHEET_NAME)
                    ws.append(HEADERS)
            else:
                _ensure_dir(client, disk_path)
                wb = Workbook()
                ws = wb.active
                ws.title = SHEET_NAME
                ws.append(HEADERS)

            ws.append([name, contact, description, datetime.now().strftime("%Y-%m-%d %H:%M"), "сайт"])
            wb.save(lp)
            _upload_retry(client, lp, disk_path)

        logger.info("Заявка с сайта сохранена: %s", name)
        return True
    except Exception as e:
        logger.exception("Ошибка сохранения заявки: %s", e)
        return False
